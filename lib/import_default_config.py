#!/usr/bin/python3

import os
import sys
import inspect

import openpyxl
import pymysql

from apps.models.config import Config

currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.insert(0, parentdir)
from settings import *

conn = pymysql.connect(
    host=MYSQL_HOST,
    port=MYSQL_PORT,
    passwd=MYSQL_PWD,
    db=MYSQL_DB,
    user=MYSQL_USER,
    charset='utf8'
)
cur = conn.cursor(cursor=pymysql.cursors.DictCursor)

# 读取xlsx方法
def import_config(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    data = []
    for row in sheet.rows:
        for cell in row:
            if cell.value != None:
                value = int(cell.value)
                data.append(value)
    return data


# 上传新的xlsx/xls文件之后(半)自动更新default_config表
def import_default_config():
    default_array = []
    level_dict = {}
    path = r"/home/web/tornado_server/game_config/default_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/default_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/default_config.xlsx"

    # workbook = openpyxl.load_workbook()
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    n = 0
    data = []
    for row in sheet.rows:
        for cell in row:
            if cell.value != None:
                value = int(cell.value)
                data.append(value)

        n += 1
        if len(data) == 100:
            default_array.append(data)
            data = []
        if n == 11:
            n = 0

    index = 0
    for da in default_array:
        daa = [da[i:i + 10] for i in range(0, len(da), 10)]
        index += 1
        level = (index-1) * 10 if index - 1 else 1
        level_dict[level] = daa
    Config.default_config = level_dict


# 上传抽奖表
def import_loto_config():
    path = r"/home/web/tornado_server/game_config/loto_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/loto_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/loto_config.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, config = [], {}
    for row in sheet.rows:
        rows.append(row)

    for row in rows[1:]:
        value = [v.value for v in row[:-1]]
        # value = tuple(value)
        config[value[0]] = {
            "tool_id": value[1],
            "num": value[2],
            "per": value[3]
        }
    Config.loto_config = config


def import_sign_config():
    path = r"/home/web/tornado_server/game_config/sign_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/sign_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/sign_config.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, sign_dict = [], {}
    for row in sheet.rows:
        rows.append(row)

    # 从第二行开始读
    for row in rows[1:]:
        value = [v.value for v in row]
        sign_dict[value[0]] = value[1]
    Config.sign_config = sign_dict


def import_sign_award_config():
    path = r"/home/web/tornado_server/game_config/sign_award_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/sign_award_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/sign_award_config.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, sign_award_dict = [], {}
    for row in sheet.rows:
        rows.append(row)

    for row in rows[1:]:
        value = [v.value for v in row]
        # print(type(value[1]))
        day_dict = {}
        value_1 = value[1].split(',')
        for v1 in value_1:
            v1_list = v1.split(':')
            # v1[0]: v1[1]
            day_dict[v1_list[0]] = v1_list[1]
        sign_award_dict[value[0]] = day_dict
    Config.sign_award_config = sign_award_dict


def import_tool_config():
    path = r"/home/web/tornado_server/game_config/tool_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/tool_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/tool_config.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, tools_dict = [], {1: {}}
    for row in sheet.rows:
        rows.append(row)

    for row in rows[1:]:
        value = [v.value for v in row]
        tools_dict[value[0]] = {
            "name": value[1],
            "source_id": value[2],
            "price": value[3],
            "describe": value[4]
        }

    Config.tool_config = tools_dict


def import_robot_config():
    path = r"/home/web/tornado_server/game_config/robot_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/robot_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/robot_config.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, robot_dict = [], {}
    for row in sheet.rows:
        rows.append(row)

    for row in rows[1:]:
        value = [v.value for v in row]
        robot_dict[value[0]] = {
            "id": value[0],
            "name": value[1],
            "avatar": value[2]
        }
    Config.robot_config = robot_dict


def import_challenge_config():
    path = r"/home/web/tornado_server/game_config/challenge_config.xlsx"
    # path = r"/mnt/d/tornado_server/game_config/challenge_config.xlsx"
    # path = r"/Users/zlinxx/tornado_server/game_config/challenge_config.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows, challenge_dict = [], {}
    for row in sheet.rows:
        rows.append(row)

    for row in rows[2:]:
        value = [v.value for v in row]
        level = value[0]
        red = value[1]
        blue = value[2]
        green = value[3]
        yellow = value[4]
        biolet = value[5]
        score = value[6]
        step = value[7]
        sec = value[8]
        star3 = str(value[9])
        star2 = str(value[10])
        star1 = str(value[11])

        if level:
            challenge_dict[level] = {}
        # if red:
        #     challenge_dict[level]["red"] = red
        # if blue:
        #     challenge_dict[level]["blue"] = blue
        # if green:
        #     challenge_dict[level]["green"] = green
        # if yellow:
        #     challenge_dict[level]["yellow"] = yellow
        # if biolet:
        #     challenge_dict[level]["biolet"] = biolet
        # if score:
        #     challenge_dict[level]["score"] = score
        # if step:
        #     challenge_dict[level]["step"] = step
        # if sec:
        #     challenge_dict[level]["sec"] = sec

        challenge_dict[level]["red"] = red
        challenge_dict[level]["blue"] = blue
        challenge_dict[level]["green"] = green
        challenge_dict[level]["yellow"] = yellow
        challenge_dict[level]["biolet"] = biolet
        challenge_dict[level]["score"] = score
        challenge_dict[level]["step"] = step
        challenge_dict[level]["sec"] = sec
        challenge_dict[level]["star3"] = star3
        challenge_dict[level]["star2"] = star2
        challenge_dict[level]["star1"] = star1

    Config.challenge_config = challenge_dict


if __name__ == "__main__":
    import_robot_config()
